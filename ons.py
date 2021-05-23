import re
from argparse import ArgumentParser
from datetime import datetime, date
from decimal import Decimal
from pathlib import Path
from typing import Iterable, Mapping

from openpyxl import load_workbook

from download import write_csv

Processing = Mapping[callable, Iterable[str]]


def pct_empty(row):
    total = len(row)
    none = sum(1 if e is None else 0 for e in row)
    return none/total


def ignore_empty(rows, threshold):
    for row in rows:
        if pct_empty(row) < threshold:
            yield row


def worksheet_rows(path, sheet):
    wb = load_workbook(path, read_only=True)
    # FFS :-/
    stripped = {n.strip(): n for n in wb.sheetnames}
    return wb[stripped[sheet]].values


def confidence(base, header_rows):
    levels = next(header_rows)
    if len(levels) == 1:
        parts = levels[0].split()
        assert parts[0] == '95%', parts
        yield f'{base}-{parts[1].lower()}-95'
        levels = next(header_rows)
        parts = levels[0].split()
        assert parts[0] == '95%', parts
        yield f'{base}-{parts[1].lower()}-95'
        return
    assert levels[0].startswith('95%'), levels[0]
    yield f'{base}-{levels[1].lower()}-95'
    levels = next(header_rows)
    assert levels[0] is None, levels[0]
    yield f'{base}-{levels[1].lower()}-95'


def extract_headers(rows, rename, header_processing):
    headers = []
    header_row = next(rows)
    if any((isinstance(h, str) and 'Lower' in h) for h in header_row):
        header_rows = iter([[h] for h in header_row])
    else:
        header_rows = zip(*(header_row, next(rows)))
    for levels in header_rows:
        if not any(levels):
            break
        base = (levels[0] or '').strip()
        base = rename.get(base, base)
        headers.append(base)
        processor = header_processing.get(base)
        if processor:
            headers.extend(processor(base, header_rows))
    return headers


def extract_period(row, name):
    period = row.pop(name).split(' to ')
    row['start'], row['end'] = (datetime.strptime(ds.strip(), '%d %B %Y').date()
                                for ds in period)
    row['mid'] = row['start'] + (row['end']-row['start'])/2


def fix_pct(row, name):
    if row['end'] < date(2020, 7, 6):
        row[name] = row[name] / 100


def datetime_to_date(row, name):
    row[name] = row[name].date()


def extract_dicts(rows, headers, row_processing):
    for row in rows:
        row = dict(zip(headers, row))
        try:
            for key, processor in row_processing.items():
                processor(row, key)
        except:
            print(row)
            raise
        yield row


def invert(mapping: Processing) -> Mapping[str, callable]:
    result = {}
    for callable_, names in mapping.items():
        for name in names:
            result[name] = callable_
    return result


def process(path: Path, sheet: str, filename: str, rename: Mapping[str, str],
            header_processing: Processing, row_processing: Processing):
    rows = worksheet_rows(path, sheet)
    rows = ignore_empty(rows, threshold=0.85)
    headers = extract_headers(rows, rename, invert(header_processing))
    dicts = extract_dicts(rows, headers, invert(row_processing))
    write_csv(dicts, filename)


def main():
    parser = ArgumentParser()
    parser.add_argument('path', help='xlsx file path', type=Path)
    args = parser.parse_args()

    date_text = re.match(r'covid19infectionsurveydatasets(\d{8})', args.path.stem).group(1)
    date = datetime.strptime(date_text, '%Y%m%d').date()
    process(
        args.path, '1a', f'ons_weekly_england_{date}.csv',
        rename={
            'Estimated average % of the population that had COVID-19': 'percent',
            'Estimate of the number of people testing positive for COVID-19': 'number',
            'Estimated number of people testing positive for COVID-19': 'number',
            'Estimated average ratio of the population that had COVID-19': 'ratio',
        },
        header_processing={confidence: ['percent', 'number', 'ratio']},
        row_processing={
            extract_period: ['Time period'],
            fix_pct: ['percent', 'percent-lower-95', 'percent-upper-95']
        }
    )
    process(
        args.path, '1b', f'ons_daily_england_{date}.csv',
        rename={
            'Modelled % testing positive for COVID-19': 'percent',
            'Modelled estimate of the number of people testing positive for COVID-19': 'number',
            'Modelled number of people testing positive for COVID-19': 'number',
            'Ratio of estimated number of people testing positive for COVID-19': 'ratio',
            'Modelled ratio of people testing positive for COVID-19': 'ratio',
        },
        header_processing={confidence: ['percent', 'number', 'ratio']},
        row_processing={datetime_to_date: ['Date']},
    )


if __name__ == '__main__':
    main()
